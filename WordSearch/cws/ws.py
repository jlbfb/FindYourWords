import logging
from collections import defaultdict, deque
from random import randint
from openpyxl import Workbook
from openpyxl.styles import Alignment
from .ws_config import directions, diff_dir, diff_fill
from django.conf import settings


# This retrieves a Python logging instance (or creates it)
logger = logging.getLogger(__name__)


def timer(fn):
    from time import perf_counter

    def calculate(*args, **kwargs):
        start = perf_counter()
        call_func = fn(*args, **kwargs)
        end = perf_counter()
        elapsed = end - start
        logger.debug(f'{fn.__name__} took {elapsed}')
        return call_func
    return calculate


def sort_available_words(available_words):
    for word in available_words:
        word[2] = len(word[1])
    available_words = sorted(available_words, key=lambda x: x[2])
    # logger.debug(f'Sorted: {available_words}')
    return available_words, available_words[0]


def build_available_words(potential_placements):
    available_words = list()
    for word, spaces in potential_placements.items():
        start_options = list()
        for space in spaces:
            start_options.append(space)
        available_words.append([word, start_options, len(spaces)])
    return available_words


def word_collector(word_list: list):
    """
    Captures the user's word collection
    """
    words = sorted(word_list, key=lambda x: len(x))
    words.reverse()
    logger.debug(f'Words sorted by length: {words}')
    logger.debug(f'Word List: {word_list}')
    word_set = dict()
    # count = 0
    for count in range(0, len(words)):
        word_set[count] = {'word': words[count], 'set': {*words[count]}}
        # count += 1
    logger.debug(f'Word Set: {word_set}')
    return word_set


def grid_builder(grid: int):
    """
    Builds the grid map based upon the selected grid size
    """
    # global grid_map
    grid_map = dict()
    for i in range(0, grid):
        for j in range(0, grid):
            grid_map[f'{i}-{j}'] = '.'
    logger.debug('Grid built')
    return grid_map


def diff_and_dir(difficulty, word_set):
    """
    Set word direction based upon the chosen difficulty level
    """
    randomizer = str()
    # logger.debug(diff_dir)
    for d, p in diff_dir[difficulty]:
        randomizer = randomizer + (str(d) * int(p * 100))
        # logger.debug(randomizer)
    for count in range(0, len(word_set)):
        rand_int = randint(0, 99)
        # logger.debug(rand_int)
        dir_select = randomizer[rand_int]
        word_set[count]['dir'] = directions[int(dir_select)]
        # logger.debug(rand_int, dir_select, word_set[count]['dir'])
    return word_set


@timer
def generate_start_positions(word_set, grid_size):
    """
    Determine all possible starting spots for all words
    """
    space_options = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    potential_placements = defaultdict(lambda: defaultdict(list))
    count = 0
    while count < len(word_set):
        this_word = word_set[count]['word']
        word_length = len(this_word)
        x_dir = word_set[count]['dir'][1]
        y_dir = word_set[count]['dir'][2]
        x_step = x_dir
        y_step = y_dir
        if x_dir > 0:
            start_x = 0
            final_x = grid_size - word_length + 1
        elif x_dir < 0:
            start_x = grid_size - 1
            final_x = word_length - 2
        else:
            start_x = 0
            final_x = grid_size
            x_step = 1
        if y_dir > 0:
            start_y = 0
            final_y = grid_size - word_length + 1
        elif y_dir < 0:
            start_y = grid_size - 1
            final_y = word_length - 2
        else:
            start_y = 0
            final_y = grid_size
            y_step = 1
        # Build the grid space options dictionary
        x_count = 0
        y_count = 0
        start_space = ''
        for row in range(start_x, final_x, x_step):
            for col in range(start_y, final_y, y_step):
                letter_count = 0
                while letter_count <= word_length:
                    current_space = f'{row + x_count}-{col + y_count}'
                    if not start_space:
                        start_space = current_space
                    potential_placements[this_word][start_space].append(
                        [this_word[letter_count], current_space]
                    )
                    space_options[
                        current_space][
                            f'{this_word[letter_count]}'][
                                this_word].append(start_space)
                    if letter_count < word_length - 1:
                        letter_count += 1
                    else:
                        start_space = ''
                        x_count = 0
                        y_count = 0
                        break
                    x_count = x_count + x_dir
                    y_count = y_count + y_dir
                col = col + y_dir
            row = row + x_dir
        count += 1
    return space_options, potential_placements


@timer
def new_word_placer(
    space_options,
    potential_placements,
    grid_map,
    word_set,
    grid_size,
    difficulty,
):
    """
    Take words in the order of fewest options and place them on the board.
    If a word has no options, move the previously placed word to a new loca-
    tion on the board. Continue removing and re-placing words until the board
    is filled or all placement options have been exhausted. Once the starting
    word has exhausted all its options, the space options are reset and place-
    ment is attempted again (max retries: two).
    """
    efforts = 0
    available_words = list()
    placed_words = defaultdict(lambda: defaultdict(list))
    placement_order = deque()
    available_words = build_available_words(potential_placements)
    while available_words:
        available_words, the_word = sort_available_words(available_words)
        if not the_word[2]:
            while True:
                logger.info(
                    f'***** The word {the_word[0]} has no place to go *****'
                )
                last_word = placement_order.pop()
                logger.debug(f'Last Word Placed: {last_word}')
                available_words.append([
                    last_word,
                    placed_words[last_word]['blocked'][last_word],
                    0
                ])
                for index, word in enumerate(available_words):
                    if word[0] == last_word:
                        continue
                    elif placed_words[last_word]['blocked'][word[0]]:
                        for blocked in (
                            placed_words[last_word]['blocked'][word[0]]
                        ):
                            available_words[index][1].append(
                                blocked
                            )
                available_words, the_word = (
                    sort_available_words(available_words)
                )
                last_word_spaces = placed_words[last_word]['spaces'][0]
                placed_words.pop(last_word)
                # logger.debug(f'Placed Words after Pop: {placed_words}')

                # Clear the last word from the board and identify any shared
                # spaces; leave the existing letter if shared
                for letter in last_word_spaces:
                    keep_letter = 0
                    for word in placed_words:
                        for word_spaces in placed_words[word]['spaces'][0]:
                            if letter[1] == word_spaces[1]:
                                logger.debug(
                                    f'Keeping Space: {word_spaces[1]} '
                                    f'for Word: {word}'
                                )
                                keep_letter = 1
                                break
                        if keep_letter:
                            break
                    if not keep_letter:
                        grid_map[letter[1]] = '.'
                # logger.debug(f'Grid Map now: {grid_map}')

                if not the_word[2]:
                    if not placement_order:
                        efforts += 1
                        if efforts < 4:
                            logger.info(
                                f'******* After failed try #{efforts}, '
                                'reshuffling word placement options *******'
                            )
                            # Reset word positions based upon difficulty
                            # after two attempts
                            if efforts == 2:
                                logger.info(
                                    f'******* After failed try #{efforts}, '
                                    'also changing word directions *******'
                                )
                                word_set = diff_and_dir(difficulty, word_set)

                            space_options, potential_placements = (
                                generate_start_positions(word_set, grid_size)
                            )
                            available_words = build_available_words(
                                potential_placements
                            )
                            available_words, the_word = (
                                sort_available_words(available_words)
                            )
                            break
                        else:
                            grid_map = 'CannotBuild'
                            logger.info(
                                'Cannot build a board for this word list '
                                f'after {efforts} attempts'
                            )
                            return grid_map
                else:
                    break

        random_index = randint(0, the_word[2]-1)
        random_start = available_words[0][1][random_index]
        used_spaces = potential_placements[the_word[0]][random_start]
        logger.debug(
            f'The word: {the_word[0]} starting at {random_start}: '
            f'Used Spaces = {used_spaces}'
        )
        placement_order.append(the_word[0])
        logger.debug(f'Placement Order: {placement_order}')
        placed_words[the_word[0]]['spaces'].append(used_spaces)
        placed_words[the_word[0]]['start'].append(random_start)
        placed_words[the_word[0]]['blocked'] = defaultdict(list)
        available_words[0][1].remove(random_start)
        placed_words[the_word[0]]['blocked'][the_word[0]] = (
            available_words.pop(0)[1]
        )
        # logger.debug(
        #     f'The Word = {the_word}\nAvailable Words = {available_words}'
        # )
        for letter in used_spaces:
            grid_map[letter[1]] = [letter[0], 1]
            for index, word in enumerate(available_words):
                for poss_letter in space_options[letter[1]]:
                    if poss_letter != letter[0]:
                        for st_space in (
                            space_options[letter[1]][poss_letter][word[0]]
                        ):
                            if st_space in word[1]:
                                available_words[index][1].remove(st_space)
                                placed_words[
                                    the_word[0]][
                                        'blocked'][
                                            word[0]].append(st_space)

        # logger.debug(f'Grid Map: {grid_map}')
        # logger.debug(f'Placed Words = {placed_words}')

    logger.info(
        f'All words have been placed (retries: {efforts}); '
        'time to build the grid!'
    )
    return grid_map


@timer
def grid_filler(word_set, grid, difficulty, grid_map):
    word_chars = str()
    for count in range(0, len(word_set)):
        word_chars = word_chars + str(word_set[count]['word'])
    for i in range(0, grid):
        for j in range(0, grid):
            # global grid_map
            char_picker = randint(0, 99)
            char_picker = ('alpha' if char_picker < diff_fill[difficulty][0]
                           else 'words')
            # logger.debug(char_picker, word_chars)
            rand_char = (randint(0, len(word_chars)-1)
                         if char_picker == 'words'
                         else randint(65, 90))
            # logger.debug(rand_char)
            rand_char = (chr(rand_char) if char_picker == 'alpha'
                         else word_chars[rand_char].upper())
            try:
                if grid_map[f'{i}-{j}'][1] == 1:
                    continue
            except Exception:
                grid_map[f'{i}-{j}'] = (rand_char, 0)
    return grid_map


def grid_map_display(grid, grid_map):
    for i in range(0, grid):
        logger.debug('\n')
        for j in range(0, grid):
            logger.debug(grid_map[f'{i}-{j}'][0], end='')
    logger.debug('\n')


def grid_map_for_template(grid, grid_map):
    grid_map_template = []
    grid_map_template.append('<table><tbody>')
    for i in range(0, grid):
        grid_map_template.append('<tr>')
        for j in range(0, grid):
            grid_map_template.append(
                f'<td class="grid">{grid_map[f"{i}-{j}"][0]}</td>'
            )
        grid_map_template.append('</tr>')
    grid_map_template.append('</tbody></table>')
    return grid_map_template


@timer
def grid_map_export_txt(
    grid, grid_map, word_collection, wc2, word_list, key='N'
):
    if key == 'Y':
        text_file = f'{settings.MEDIA_ROOT}/grids/{wc2}_Key.txt'
    else:
        text_file = f'{settings.MEDIA_ROOT}/grids/{wc2}.txt'
    with open(text_file, 'w') as f:
        f.write(f'Word Search\n{word_collection}\n\n')
        for i in range(0, grid):
            f.write('\n')
            for j in range(0, grid):
                if grid_map[f'{i}-{j}'][1] == 1:
                    f.write(grid_map[f'{i}-{j}'][0])
                else:
                    if key == 'Y':
                        f.write('.')
                    else:
                        f.write(grid_map[f'{i}-{j}'][0])
                f.write(' ')
        f.write('\n\n')
        for word in word_list:
            f.write(f"{word}\n")
    logger.debug('Done')


@timer
def grid_map_export_excel(
    grid, grid_map, word_collection, wc2, word_list, key='N'
):
    cell_ctr = Alignment(horizontal='center')
    wb = Workbook()
    ws = wb.active
    ws.title = word_collection
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=grid
    )
    ws.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=grid
    )
    ws.merge_cells(
        start_row=3, start_column=1, end_row=3, end_column=grid
    )
    ws.cell(row=1, column=1).value = 'Word Search'
    ws.cell(row=1, column=1).alignment = cell_ctr
    ws.cell(row=2, column=1).value = word_collection
    ws.cell(row=2, column=1).alignment = cell_ctr
    curr_row = 3
    for i in range(0, grid):
        curr_col = 1
        curr_row += 1
        for j in range(0, grid):
            ws.column_dimensions[chr(curr_col + 64)].width = 2.5
            if grid_map[f'{i}-{j}'][1] == 1:
                ws.cell(row=curr_row, column=curr_col
                        ).value = grid_map[f'{i}-{j}'][0]
            else:
                if key == 'Y':
                    ws.cell(row=curr_row, column=curr_col).value = '.'
                else:
                    ws.cell(row=curr_row, column=curr_col
                            ).value = grid_map[f'{i}-{j}'][0]
            ws.cell(row=curr_row, column=curr_col
                    ).alignment = cell_ctr
            curr_col += 1
    curr_row += 1
    ws.merge_cells(start_row=curr_row, start_column=1,
                   end_row=curr_row, end_column=grid)
    for word in word_list:
        curr_row += 1
        ws.cell(row=curr_row, column=1
                ).value = word
    # TODO: Add Try/ Except handling to cover file being open with same name
    if key == 'Y':
        wb.save(f'{settings.MEDIA_ROOT}/grids/{wc2}_Key.xlsx')
    else:
        wb.save(f'{settings.MEDIA_ROOT}/grids/{wc2}.xlsx')
    logger.debug('Done')


# @timer
# def save_collections():
#     global this_key
#     global word_collection
#     while True:
#         try:
#             tw.insert_table('Word_Collections',
#                 word_collection = (word_collection,))
#             this_key = tw.get_key('Word_Collections', id_ = 'id',
#                 identifier = 'word_collection', value = word_collection)[0]
#             for w in range(0, len(word_set)):
#                 if w == len(word_set) - 1:
#                     loop = 'last'
#                 else:
#                     loop = 'next'
#                 tw.insert_table('Words', word_collections_id = this_key,
#                     word = word_set[w]['word'], loop = loop)
#             break
#         except Exception as e:
#             logger.debug(e)
#             word_collection = input('Please enter another name for the '
#                 'collection: ')


# @timer
# def save_grid_maps():
#     this_grid = tw.get_key('Grid_Maps', id_ = 'grid_maps_id')
#     if this_grid == []:
#         this_grid = 1
#     else:
#         this_grid = max(this_grid)
#         logger.debug(this_grid)
#         this_grid = this_grid[0] + 1
#     logger.debug(this_grid)
#     for i in range(0, grid):
#         for j in range(0, grid):
#             if j == grid - 1:
#                 loop = 'last'
#             else:
#                 loop = 'next'
#             tw.insert_table(
#                 'Grid_Maps', loop=loop, grid_maps_id=this_grid,
#                 cell=f'{i}-{j}', letter=grid_map[f'{i}-{j}'][0],
#                 word_key=grid_map[f'{i}-{j}'][1]
#             )
#     tw.insert_table('Collection_Grid', word_collections_id = this_key,
#         grid_maps_id = this_grid, difficulty = difficulty, grid_size = grid)


# def load_collections():
#     global word_list
#     word_list = []
#     collection_list = tw.get_data('Word_Collections', columns = '*')
#     logger.debug('Available collections:')
#     for collection in collection_list:
#         logger.debug(f'{collection[0]}: {collection[1]}')
#     while True:
#         collection = input('Select the list to load: ')
#         if re.match('^[0-9]{1,2}$', collection):
#             try:
#                 collection = int(collection)
#                 words = tw.get_data('Words', columns = '*',
#                     condition = f"word_collections_id = \'{collection}\'")
#                 logger.debug(
#                     f'Words in the {collection_list[collection - 1][1]} '
#                     'Collection:'
#                 )
#                 for word in words:
#                     word_list.append(word[2])
#                     logger.debug(word[2])
#                 break
#             except Exception as e:
#                 logger.debug(f'{e}\nThat is not a valid selection.')
#         else:
#             logger.debug('That is not a valid selection.')
#     change_word()
#     if changed_words != []:
#         for change in changed_words:
#             tw.update_data('Words', word = word_list[change],
#                 id = words[change][0])
#             logger.debug(
#                 f'Replaced {words[change][0]} with {word_list[change]}'
#             )
#     else:
#         logger.debug('No changes made.')
